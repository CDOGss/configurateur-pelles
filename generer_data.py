# -*- coding: utf-8 -*-
"""
generer_data.py
================
Régénère le fichier `data.js` du configurateur à partir de DEUX fichiers Excel :

    1. le STOCK d'attaches   (par défaut : Lesstocks.xlsx)
    2. la liste des MACHINES (par défaut : Machines.xlsx)

Utilisation
-----------
    python generer_data.py
    python generer_data.py MonStock.xlsx MesMachines.xlsx

Des modèles à respecter sont fournis dans le dossier `modeles/`.
Il suffit de garder les mêmes en-têtes, puis de relancer ce script :
`data.js` est réécrit et l'application est à jour (rechargez la page).

Format attendu
--------------
STOCK    : ligne 1 = les en-têtes (types d'attelage). Chaque colonne liste
           les attaches en stock de ce type, de haut en bas.
MACHINES : ligne 1 = en-têtes. Colonnes reconnues (accents/majuscules ignorés) :
           - Designation  (obligatoire)
           - Matricule    (obligatoire)
           - Tonnage      (optionnel — sinon déduit du modèle, ex. DX380 -> 38 t)
           - Type         (optionnel — "chenilles" / "pneus", sinon déduit :
                           LC -> chenilles, W/WR -> pneus)
"""

import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl est requis. Installez-le avec :  pip install openpyxl")

# Palette de couleurs par colonne (réutilisée dans l'ordre des en-têtes)
PALETTE = [
    "#2f80ed", "#9b51e0", "#eb5757", "#e2892b", "#17a398", "#2d9cdb",
    "#6a994e", "#7a6ff0", "#e0a800", "#c2185b", "#00897b", "#5c6bc0",
]


# ----------------------------------------------------------------------
# Utilitaires
# ----------------------------------------------------------------------
def sans_accent(txt):
    """minuscule + sans accent + sans espaces superflus (pour comparer)."""
    txt = unicodedata.normalize("NFD", str(txt))
    txt = "".join(c for c in txt if unicodedata.category(c) != "Mn")
    return txt.strip().lower()


def normaliser_espaces(txt):
    """réduit les espaces multiples à un seul."""
    return re.sub(r"\s+", " ", str(txt)).strip()


def id_colonne(entete):
    """id technique pour une colonne (sert de clé) : lettres/chiffres seulement."""
    return re.sub(r"[^A-Za-z0-9]", "", str(entete)) or "COL"


def joli_titre(entete):
    """'DIAM65' -> 'DIAM 65' ; le reste inchangé."""
    m = re.match(r"^(DIAM)(\d+)$", str(entete).strip(), re.IGNORECASE)
    return f"{m.group(1).upper()} {m.group(2)}" if m else str(entete).strip()


def deduire_tonnage(designation):
    """DX380 -> 38.0 ; DX165 -> 16.5 (nombre du modèle / 10)."""
    m = re.search(r"DX\s*0*(\d{2,4})", designation, re.IGNORECASE)
    if not m:
        return None
    val = int(m.group(1)) / 10.0
    return int(val) if val == int(val) else val


def deduire_type(designation):
    """chargeuse / LC(R) -> chenilles / W(R) -> pneus ; sinon d'après le libellé."""
    s = sans_accent(designation)
    # chargeuse sur pneus (silhouette différente d'une pelle)
    if "chargeuse" in s or "loader" in s or re.search(r"\bdl\s*\d+", s):
        return "chargeuse"
    if "chenille" in s:
        return "chenilles"
    if "pneu" in s:
        return "pneus"
    # Doosan/Develon : DX###LC / LCR -> chenilles ; DX###W / WR -> pneus
    m = re.search(r"dx\s*\d+\s*(lcr|lc|wr|w)", s)
    if m:
        return "chenilles" if m.group(1).startswith("lc") else "pneus"
    # Liebherr : A### -> sur pneus ; R### -> sur chenilles
    if re.search(r"\ba\s*\d{3}\b", s):
        return "pneus"
    if re.search(r"\br\s*\d{3}\b", s):
        return "chenilles"
    return None


# ----------------------------------------------------------------------
# Lecture des fichiers
# ----------------------------------------------------------------------
def lire_stock(chemin):
    """Retourne (colonnes, stock) à partir du fichier de stock."""
    ws = load_workbook(chemin, data_only=True).active
    lignes = list(ws.iter_rows(values_only=True))
    if not lignes:
        sys.exit(f"[stock] '{chemin}' est vide.")

    entetes = lignes[0]
    colonnes, stock, ids_vus = [], {}, set()

    for i, entete in enumerate(entetes):
        if entete is None or str(entete).strip() == "":
            continue
        cid = id_colonne(entete)
        while cid in ids_vus:          # garantit des id uniques
            cid += "_"
        ids_vus.add(cid)
        colonnes.append({
            "id": cid,
            "titre": joli_titre(entete),
            "couleur": PALETTE[len(colonnes) % len(PALETTE)],
        })
        # valeurs de la colonne (on ignore les cellules vides)
        valeurs = []
        for ligne in lignes[1:]:
            cell = ligne[i] if i < len(ligne) else None
            if cell is not None and str(cell).strip() != "":
                valeurs.append(normaliser_espaces(cell))
        stock[cid] = valeurs

    if not colonnes:
        sys.exit(f"[stock] Aucun en-tête trouvé en ligne 1 de '{chemin}'.")
    return colonnes, stock


def lire_machines(chemin):
    """Retourne la liste des machines à partir du fichier machines."""
    ws = load_workbook(chemin, data_only=True).active
    lignes = list(ws.iter_rows(values_only=True))
    if not lignes:
        sys.exit(f"[machines] '{chemin}' est vide.")

    # repérage des colonnes par nom (accents/casse ignorés)
    entetes = [sans_accent(x) if x is not None else "" for x in lignes[0]]

    def idx(*noms):
        for n in noms:
            if n in entetes:
                return entetes.index(n)
        return None

    i_des = idx("designation", "designations", "modele", "libelle")
    i_mat = idx("matricule", "mat", "num", "numero")
    i_ton = idx("tonnage", "tonnes", "poids")
    i_typ = idx("type", "roulement")

    if i_des is None or i_mat is None:
        sys.exit("[machines] Il faut au minimum les colonnes 'Designation' et 'Matricule'.")

    machines, avertissements = [], []
    n = 0
    for ligne in lignes[1:]:
        des = ligne[i_des] if i_des < len(ligne) else None
        if des is None or str(des).strip() == "":
            continue
        n += 1
        designation = normaliser_espaces(des)
        matricule = normaliser_espaces(ligne[i_mat]) if i_mat < len(ligne) and ligne[i_mat] is not None else "????"

        # tonnage
        tonnage = None
        if i_ton is not None and i_ton < len(ligne) and ligne[i_ton] not in (None, ""):
            try:
                tonnage = float(str(ligne[i_ton]).replace(",", "."))
                tonnage = int(tonnage) if tonnage == int(tonnage) else tonnage
            except ValueError:
                tonnage = None
        if tonnage is None:
            tonnage = deduire_tonnage(designation)
        if tonnage is None:
            tonnage = 21
            avertissements.append(f"  • {designation} : tonnage indéterminé -> 21 t par défaut")

        # type
        typ = None
        if i_typ is not None and i_typ < len(ligne) and ligne[i_typ] not in (None, ""):
            t = sans_accent(ligne[i_typ])
            if "chenille" in t or "crawler" in t:
                typ = "chenilles"
            elif "pneu" in t or "wheel" in t or "roue" in t:
                typ = "pneus"
        if typ is None:
            typ = deduire_type(designation)
        if typ is None:
            typ = "chenilles"
            avertissements.append(f"  • {designation} : type indéterminé -> chenilles par défaut")

        machines.append({
            "id": f"m{n}",
            "designation": designation,
            "matricule": matricule,
            "tonnage": tonnage,
            "type": typ,
        })

    if not machines:
        sys.exit("[machines] Aucune machine trouvée.")
    return machines, avertissements


# ----------------------------------------------------------------------
# Écriture de data.js
# ----------------------------------------------------------------------
def ecrire_data_js(colonnes, stock, machines, chemin_sortie):
    j = lambda o: json.dumps(o, ensure_ascii=False, indent=2)
    contenu = f"""/* =============================================================
   data.js — GÉNÉRÉ AUTOMATIQUEMENT par generer_data.py
   Ne pas éditer à la main : relancez le script pour mettre à jour.
   ============================================================= */

// En-têtes des colonnes (types d'attelage), dans l'ordre du fichier
const COLONNES = {j(colonnes)};

// Attaches en stock, par colonne
const STOCK = {j(stock)};

// Les machines (id, désignation, matricule, tonnage, type)
const MACHINES = {j(machines)};
"""
    Path(chemin_sortie).write_text(contenu, encoding="utf-8")


# ----------------------------------------------------------------------
# Programme principal
# ----------------------------------------------------------------------
def main():
    args = sys.argv[1:]
    fichier_stock = args[0] if len(args) >= 1 else "Lesstocks.xlsx"
    fichier_machines = args[1] if len(args) >= 2 else "Machines.xlsx"

    for f in (fichier_stock, fichier_machines):
        if not Path(f).exists():
            sys.exit(f"Fichier introuvable : {f}\n"
                     f"Astuce : dupliquez un modèle du dossier 'modeles/' et remplissez-le.")

    print(f"Lecture du stock    : {fichier_stock}")
    colonnes, stock = lire_stock(fichier_stock)
    print(f"Lecture des machines: {fichier_machines}")
    machines, avertissements = lire_machines(fichier_machines)

    ecrire_data_js(colonnes, stock, machines, "data.js")

    total_attaches = sum(len(v) for v in stock.values())
    print("\n data.js régénéré avec succès.")
    print(f"   {len(colonnes)} colonne(s) : " + ", ".join(c["titre"] for c in colonnes))
    print(f"   {total_attaches} attache(s) au total")
    print(f"   {len(machines)} machine(s) : " + ", ".join(m["matricule"] for m in machines))
    if avertissements:
        print("\n Avertissements :")
        print("\n".join(avertissements))
    print("\n Rechargez l'application dans le navigateur pour voir la mise à jour.")


if __name__ == "__main__":
    main()
